#!/usr/bin/python
import openpyxl
import os.path
import os 
import sys
from psycopg2 import sql
from glob import glob
from datetime import datetime
from config import connect_postgres
from config import connect_oracle

date=datetime.today().strftime('%d/%m/%Y') 
    
#Read the excel file
def read_file():
    try:
        path = glob(os.path.join("*.xlsx"))[0]        
        wb_obj = openpyxl.load_workbook(path)         
        sheet_obj = wb_obj.active 
        #excel column names:
        column_user = 'usuario'
        column_ehumano = 'ehumano'        
        col_user=[]
        col_ehumano=[]
        
        for column_cell in sheet_obj.iter_cols(1, sheet_obj.max_column): 
             
            if column_cell[0].value == column_user:    
                
                for data in column_cell[1:]:                    
                    col_user.append(data.value)                    
                break
            
            if column_cell[0].value == column_ehumano:    
                
                for data in column_cell[1:]:                     
                    col_ehumano.append(data.value)                    
                break

    except Exception as e:
        print("Archivo .xlsx no encontrado") 
        input('Presione una tecla para salir ')
        sys.exit()
        
    return col_user, col_ehumano

#Update the state by usuario or by ehumano
def update_state():    
    conn = None
    con = None
    up_state = []
    upd_state = []
    
    try:   
        conn, cur = connect_postgres()
        con, cursor = connect_oracle()        
        col_user, col_ehumano = read_file()
        
        if col_user:       
             
            up_state, state_cat = update_catalogo(col_user, 'usuario_login', conn, cur)
            print_output2(state_cat,up_state,'USUARIO', 1) 
            
            upd_state, state_user = update_ventas(col_user, con, cursor, 'username')
            print_output2(state_user, upd_state, 'USUARIO', 2) 
       
        if col_ehumano: 
              
            up_state, state_cat = update_catalogo(col_ehumano, 'ehumano', conn, cur) 
            print_output2(state_cat,up_state,'E. HUMANO', 1)  
          
            upd_state, state_user = update_ventas(col_ehumano, con, cursor, 'iniciales')
            print_output2(state_user, upd_state, 'E. HUMANO', 2) 
                   
        
        if not col_user and not col_ehumano:
            print('No se encontraron columnas usuario o ehumano') 
        
        #cur.close()  
        cursor.close()
    
    except Exception as e:
        print("Error al actualizar datos", e) 
        input('Presione una tecla para salir ')
        sys.exit()
        
    finally:
        if conn is not None: conn.close()       
        if con is not None: con.close()  
            

def update_catalogo(col_val, colu_name, conn, cur):
    
    print("\n-------------CATÁLOGO DE CUBOS--------------\n") 
    
    up_state = []
    state_cat = get_state_catalogo(col_val, colu_name, conn, cur) 
    
    for item, state in state_cat.items():
        
        if state is not None:   
            rowc = update_item_catalogo(colu_name, item, conn, cur)
            up_state.append(rowc)
        else:
            up_state.append(0)
           
    return up_state, state_cat

def get_state_catalogo(col, col_name, conn, cur):
   
    state_cat={}
    try:   
       
        for values in col:            
            
            query = sql.SQL("SELECT flag_estado FROM usuario WHERE {cn} = %s and flag_estado = 'AC'").format(cn=sql.Identifier(col_name))
            cur.execute(query,(values,))
            data = cur.fetchone()
            if data is not None:
                state_cat[values] = data[0] 
            else:
                state_cat[values] = data
                
        conn.commit()
        
    except Exception as e:          
        print('Error', e)  
        input('Presione una tecla para salir ')
        sys.exit()  
    return state_cat

#update data

def update_item_catalogo(column_name, column_val, conn, cur):
    
    affected_rows = 0
    
    if column_name == 'ehumano':
        user = get_user(column_val, conn, cur)
    else:
        user = column_val 
           
    try:  
        username=user+"1"  
        query = sql.SQL("UPDATE usuario SET flag_estado = 'IN', email_usuario = %s, usuario_login = %s WHERE {cn} = %s and flag_estado = 'AC'").format(cn=sql.Identifier(column_name))
        cur.execute(query, (date, username, column_val,))            
        
        affected_rows = cur.rowcount 
        conn.commit() 
        
    except Exception as e:          
        print('Error', e)   
        input('Presione una tecla para salir ')
        sys.exit()
    return affected_rows 

def get_user(ehumano, conn, cur):
    
    user = " "   
    try: 
        cur.execute("SELECT usuario_login FROM usuario WHERE ehumano = %s and flag_estado = 'AC'", (ehumano,)) 
        user = cur.fetchone()[0]    
        conn.commit()
       
    except Exception as e:          
        print('Error', e)  
        input('Presione una tecla para salir ')
        sys.exit()
    return user

#---------------------------------VENTAS-----------------------------------------#

def update_ventas(col_val, con, cursor ,colu_name):
    
    print("\n-------------PÁGINA DE VENTAS-------------\n")
    
    up_state = []
    state_users = get_state_ventas(col_val, con, cursor, colu_name) 
  
    for item, state in state_users.items():
        
        if state:
            if 1 in state: 
                rowc = update_item_ventas(item, con, cursor, colu_name)
                up_state.append(rowc)  
            else: up_state.append(0)         
        else:
            up_state.append(0)
         
    return up_state, state_users

def get_state_ventas(col_values, con, cursor, col_name):
   
    state_users={}
    try:          
        for values in col_values:            
            
            if col_name == 'username':
                
                query = ("SELECT estado FROM sec_usuario WHERE UPPER(username) = '%s'")
                cursor.execute(query % (values.upper(),))
                
            else:
                query = ("SELECT estado FROM sec_usuario WHERE INICIALES = '%s'")
                cursor.execute(query % (values,))
                
            data = cursor.fetchall()            
            state_users[values] = [item for t in data for item in t] 
          
        con.commit()
        
    except Exception as e:          
        print('Error', e)  
        input('Presione una tecla para salir ')
        sys.exit()  
    return state_users

def update_item_ventas(column_val, con, cursor, col_name):
    
    affected_rows = 0
   
    try:  
        if col_name == 'username':
        
            updatequery = "UPDATE sec_usuario SET estado = 0 WHERE UPPER(username) = '%s' AND estado = 1"            
            cursor.execute(updatequery % (column_val.upper(),)) 
                      
        else:
            updatequery = "UPDATE sec_usuario SET estado = 0 WHERE INICIALES = '%s' AND estado = 1"            
            cursor.execute(updatequery % (column_val,))     
            
        affected_rows = cursor.rowcount 
        con.commit() 
        
    except Exception as e:          
        print('Error', e)   
        input('Presione una tecla para salir ')
        sys.exit()
    return affected_rows 

def print_output2(state, newstate, name, var):
    users=[]
    new_state=[]
    state_strings=[]
    
    if var == 1:
        for u,s in state.items():
            users.append(u)
            
            if s == None: 
                state_strings.append("Inactivo")
            else:
                
                if s == 'AC' or s == '1':
                    state_strings.append("Activo") 
                    
                if s == 'IN' or s == '0':
                    state_strings.append("Inactivo") 
    else:
        for u,sta in state.items():
            users.append(u)
            
            if sta:                
                if 1 in sta:
                    state_strings.append("Activo") 
                else:
                    state_strings.append("Inactivo")         
            else:
                state_strings.append("Inactivo")
                
    for i in newstate:
        if i > 0:
            new_state.append("Inactivo")
        else:
            new_state.append("Ninguno")
                                          
    print("{:<15} {:<15} {:<10} \n".format(name,'ESTADO','CAMBIOS'))
       
    for cu,st,ns in zip(users, state_strings, new_state):
       print("{:<15} {:<15} {:<10}".format(cu,st,ns))
    
    
if __name__ == '__main__':

    update_state()
    input('Presione una tecla para salir ')
    sys.exit()
    